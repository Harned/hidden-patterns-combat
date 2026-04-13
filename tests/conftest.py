from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook


@pytest.fixture
def demo_excel_path(tmp_path: Path) -> Path:
    """Create a minimal multi-sheet workbook with two-row headers.

    Includes duplicated/incomplete header cells to emulate real competition tables.
    """
    path = tmp_path / "episodes_demo.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Общее"
    ws1.append([
        None,
        None,
        None,
        "Стойка и маневрирование самбиста (основные в эпизоде)",
        "Стойка и маневрирование самбиста (основные в эпизоде)",
        "Контакты Физического Взаимодействия (захваты, обхваты, прихваты, хваты, упоры)",
        "Выведение соперника из устойчивого положения (при выполнении n или n1)",
    ])
    ws1.append([
        "ФИО борца",
        "Технико-тактический эпизод",
        "Баллы",
        "x1",
        "x1",  # duplicate second-level header
        "k1",
        "v1",
    ])
    ws1.append(["A", 1, 2, 1, 0, 1, 0])
    ws1.append(["B", 2, 0, 0, 1, "да", "нет"])

    ws2 = wb.create_sheet("48")
    ws2.append([
        None,
        None,
        "Завершающие атаку приемы (n)",
        "Контакты Физического Взаимодействия (захваты, обхваты, прихваты, хваты, упоры)",
    ])
    ws2.append([
        "ФИО борца",
        "Баллы",
        "f1",
        "k1",
    ])
    ws2.append(["C", 1, 1, 0])

    wb.save(path)
    return path


@pytest.fixture
def cleaned_like_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "metadata__athlete_name": ["A", "B", "C"],
            "metadata__episode_attr_01": [1, 2, 3],
            "outcomes__score": [2, "0", None],
            "maneuvering__indicator_01": [1, "да", "нет"],
            "maneuvering__indicator_02": [0, 1, "x"],
            "kfv__indicator_01": [1, "yes", "no"],
            "kfv__indicator_07": [0, 1, 1],
            "kfv__indicator_13": [0, 0, 1],
            "kfv__indicator_19": [0, 1, 0],
            "kfv__indicator_25": [1, 0, 0],
            "vup__indicator_01": [0, 1, 0],
            "outcomes__finish_action_01": [1, 0, 0],
            "metadata__sheet": ["Общее", "Общее", "48"],
        }
    )


@pytest.fixture
def matrix_excel_path(tmp_path: Path) -> Path:
    path = tmp_path / "episodes_matrix.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Matrix"

    ws.append(["ФИО борца", "Иванов И.И.", None, None, None])
    ws.append(["№ эпизода", None, 1, 2, 3])
    ws.append(["Время эпизода", None, 30, 45, 20])
    ws.append(["Время паузы", None, 5, 3, 2])
    ws.append(["Баллы", None, 0, 2, 0])
    ws.append(["Правосторонняя стойка", None, 1, 0, 1])
    ws.append(["Левосторонняя стойка", None, 0, 1, 0])
    ws.append(["Захваты", None, 1, 1, 0])
    ws.append(["Хваты", None, 0, 1, 0])
    ws.append(["Обхваты", None, 0, 0, 1])
    ws.append(["Прихваты", None, 1, 0, 0])
    ws.append(["Упоры", None, 0, 0, 1])
    ws.append(["ВУП", None, 0, 1, 0])
    ws.append(["Завершающие атакующие действия", None, 0, 1, 0])

    wb.save(path)
    return path
